"""Excel import logic for Project Controls summary workbooks."""
import traceback
from decimal import Decimal, InvalidOperation
from datetime import date
import openpyxl
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from .models import (
    RockDailyRecord, RockBargePlacement, Barge,
    SandDailyRecord, SandBargePlacement,
    RecoveryActionPlan, RecoveryActionItem,
    RevetmentStation, RevetmentActivity, RevetmentDailyRecord, RevetmentDailyItem,
    ImportLog,
)
from .services import recalculate_rock_accumulatives, recalculate_sand_accumulatives


def _to_decimal(val, default=Decimal('0')):
    if val is None or val == '':
        return default
    try:
        return Decimal(str(val)).quantize(Decimal('0.01'))
    except InvalidOperation:
        return default


def _to_int(val, default=0):
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return default


def _to_date(val):
    if isinstance(val, date):
        return val
    if hasattr(val, 'date'):
        return val.date()
    try:
        from datetime import datetime
        return datetime.strptime(str(val), '%Y-%m-%d').date()
    except Exception:
        return None


def run_import(project, import_type, excel_file, user, duplicate_action='update'):
    log = ImportLog(
        project=project,
        import_type=import_type,
        imported_by=user,
        status='pending',
    )
    file_content = excel_file.read()
    log.file.save(excel_file.name, ContentFile(file_content), save=False)
    log.save()

    try:
        if import_type == 'rock_summary':
            success, failed, error = _import_rock(project, file_content, duplicate_action)
        elif import_type == 'sand_summary':
            success, failed, error = _import_sand(project, file_content, duplicate_action)
        elif import_type == 'recovery_action_plan':
            success, failed, error = _import_rap(project, file_content, duplicate_action)
        elif import_type == 'revetment_progress':
            success, failed, error = _import_revetment(project, file_content, duplicate_action, user)
        else:
            log.status = 'failed'
            log.error_message = f'Unknown import type: {import_type}'
            log.save()
            return log

        log.success_rows = success
        log.failed_rows = failed
        log.total_rows = success + failed
        log.error_message = error
        log.status = 'success' if failed == 0 else ('partial' if success > 0 else 'failed')
        log.save()
    except Exception as e:
        log.status = 'failed'
        log.error_message = traceback.format_exc()[:2000]
        log.save()
    return log


def _import_rock(project, file_content, duplicate_action):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb['Rock'] if 'Rock' in wb.sheetnames else wb.active
    # Data starts at row 7 (index 6 in 0-based), columns: A=day, B=date, C=tct_daily, D=tct_accum, E=trips, F=trucks
    # G=barge1, H=barge2, I=barge3, J=barge4, K=placed_daily, L=placed_accum, M=station, N=core_outside_daily, O=core_outside_accum, P=core_inside_accum, Q=remarks
    barges = list(Barge.objects.all())
    barge_map = {b.code: b for b in barges}
    b_codes = ['B1', 'B2', 'B3', 'B4']  # Yuttachai, Bang Sapan 2, MTP10, MTP8

    success = failed = 0
    errors = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        try:
            day_name = str(row[0] or '')
            record_date = _to_date(row[1])
            if not record_date:
                continue
            tct_daily = _to_decimal(row[2])
            tct_accum = _to_decimal(row[3])
            trips = _to_int(row[4])
            trucks = _to_int(row[5])
            barge1 = _to_decimal(row[6])
            barge2 = _to_decimal(row[7])
            barge3 = _to_decimal(row[8])
            barge4 = _to_decimal(row[9])
            placed_daily = barge1 + barge2 + barge3 + barge4
            placed_accum = _to_decimal(row[10] if len(row) > 10 else None)
            station = str(row[11] or '') if len(row) > 11 else ''
            core_outside_daily = _to_decimal(row[12] if len(row) > 12 else None)
            core_outside_accum = _to_decimal(row[13] if len(row) > 13 else None)
            core_inside_accum = _to_decimal(row[14] if len(row) > 14 else None)
            remarks = str(row[15] or '') if len(row) > 15 else ''

            defaults = dict(
                day_name=day_name[:10],
                material_type=RockDailyRecord.MATERIAL_CORE,
                tct_daily_ton=tct_daily, tct_accum_ton=tct_accum,
                tct_trips=trips, tct_trucks=trucks,
                placed_daily_ton=placed_daily, placed_accum_ton=placed_accum,
                station_of_core=station[:300],
                core_outside_daily=core_outside_daily,
                core_outside_accum=core_outside_accum,
                core_inside_accum=core_inside_accum,
                remarks=remarks[:500],
            )

            if duplicate_action == 'update':
                rec, _ = RockDailyRecord.objects.update_or_create(
                    project=project, record_date=record_date, defaults=defaults)
            else:
                rec, created = RockDailyRecord.objects.get_or_create(
                    project=project, record_date=record_date, defaults=defaults)
                if not created:
                    success += 1
                    continue

            # Barge placements
            barge_qtys = [barge1, barge2, barge3, barge4]
            for code, qty in zip(b_codes, barge_qtys):
                if qty > 0 and code in barge_map:
                    RockBargePlacement.objects.update_or_create(
                        record=rec, barge=barge_map[code],
                        defaults={'quantity_ton': qty, 'trips': None},
                    )
            success += 1
        except Exception as e:
            failed += 1
            errors.append(str(e)[:100])

    if success > 0:
        recalculate_rock_accumulatives(project)
    return success, failed, '; '.join(errors[:5])


def _import_sand(project, file_content, duplicate_action):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb['Sand'] if 'Sand' in wb.sheetnames else wb.active
    barges = list(Barge.objects.all())
    barge_map = {b.code: b for b in barges}
    b_codes = ['B1', 'B2', 'B3', 'B4']

    success = failed = 0
    errors = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        try:
            day_name = str(row[0] or '')
            record_date = _to_date(row[1])
            if not record_date:
                continue
            tct_daily = _to_decimal(row[2])
            tct_accum = _to_decimal(row[3])
            tct_trips = _to_int(row[4])
            tct_trucks = _to_int(row[5])
            mtp3_daily = _to_decimal(row[6] if len(row) > 6 else None)
            mtp3_accum = _to_decimal(row[7] if len(row) > 7 else None)
            mtp3_trips = _to_int(row[8] if len(row) > 8 else None)
            mtp3_trucks = _to_int(row[9] if len(row) > 9 else None)
            chalothon_daily = _to_decimal(row[10] if len(row) > 10 else None)
            khlong_bang_phai_daily = _to_decimal(row[11] if len(row) > 11 else None)
            total_daily = tct_daily + mtp3_daily
            barge1 = _to_decimal(row[12] if len(row) > 12 else None)
            barge2 = _to_decimal(row[13] if len(row) > 13 else None)
            barge3 = _to_decimal(row[14] if len(row) > 14 else None)
            barge4 = _to_decimal(row[15] if len(row) > 15 else None)
            offshore_daily = barge1 + barge2 + barge3 + barge4
            offshore_station = str(row[16] or '') if len(row) > 16 else ''
            onshore_daily = _to_decimal(row[17] if len(row) > 17 else None)
            onshore_trips = _to_int(row[18] if len(row) > 18 else None)
            onshore_trucks = _to_int(row[19] if len(row) > 19 else None)
            inside_daily = _to_decimal(row[20] if len(row) > 20 else None)
            outside_daily = _to_decimal(row[21] if len(row) > 21 else None)
            remaining_tct = _to_decimal(row[22] if len(row) > 22 else None)
            remaining_mtp3 = _to_decimal(row[23] if len(row) > 23 else None)
            remarks = str(row[24] or '') if len(row) > 24 else ''

            defaults = dict(
                day_name=day_name[:10],
                tct_daily_ton=tct_daily, tct_accum_ton=tct_accum,
                tct_trips=tct_trips, tct_trucks=tct_trucks,
                mtp3_daily_ton=mtp3_daily, mtp3_accum_ton=mtp3_accum,
                mtp3_trips=mtp3_trips, mtp3_trucks=mtp3_trucks,
                chalothon_daily_ton=chalothon_daily,
                khlong_bang_phai_daily_ton=khlong_bang_phai_daily,
                oswald_daily_ton=khlong_bang_phai_daily,
                sand_source='Khlong Bang Phai (Oswald)',
                total_daily_ton=total_daily,
                offshore_daily_ton=offshore_daily, offshore_station=offshore_station[:300],
                onshore_daily_ton=onshore_daily,
                onshore_trips=onshore_trips, onshore_trucks=onshore_trucks,
                inside_plot_daily=inside_daily, outside_plot_daily=outside_daily,
                remaining_tct=remaining_tct, remaining_mtp3=remaining_mtp3,
                remarks=remarks[:500],
            )

            if duplicate_action == 'update':
                rec, _ = SandDailyRecord.objects.update_or_create(
                    project=project, record_date=record_date, defaults=defaults)
            else:
                rec, created = SandDailyRecord.objects.get_or_create(
                    project=project, record_date=record_date, defaults=defaults)
                if not created:
                    success += 1
                    continue

            barge_qtys = [barge1, barge2, barge3, barge4]
            for code, qty in zip(b_codes, barge_qtys):
                if qty > 0 and code in barge_map:
                    SandBargePlacement.objects.update_or_create(
                        record=rec, barge=barge_map[code],
                        defaults={'quantity_ton': qty},
                    )
            success += 1
        except Exception as e:
            failed += 1
            errors.append(str(e)[:100])

    if success > 0:
        recalculate_sand_accumulatives(project)
    return success, failed, '; '.join(errors[:5])


def _import_rap(project, file_content, duplicate_action):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb['Remaining Qty'] if 'Remaining Qty' in wb.sheetnames else wb.active

    success = failed = 0
    errors = []
    plan = None
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        plan = RecoveryActionPlan.objects.create(
            project=project,
            title=f'Recovery Action Plan – imported',
            start_date='2026-01-01',
            end_date='2026-12-31',
            status='Draft',
        )
    except Exception as e:
        return 0, 0, f'Could not create plan: {e}'

    for row in ws.iter_rows(min_row=3, values_only=True):
        try:
            if not row[0]:
                continue
            item_no = str(row[0])
            description = str(row[1] or '')[:500]
            total_qty = _to_decimal(row[2])
            unit = str(row[3] or '')[:50]
            if not description:
                continue
            RecoveryActionItem.objects.get_or_create(
                recovery_plan=plan,
                item_no=item_no,
                defaults={
                    'description': description,
                    'total_quantity': total_qty,
                    'unit': unit,
                }
            )
            success += 1
        except Exception as e:
            failed += 1
            errors.append(str(e)[:100])

    return success, failed, '; '.join(errors[:5])


def _merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _normalize_revetment_status(value):
    if value is None or value == '':
        return ''
    if _to_date(value):
        return 'Inspected'
    text = str(value).strip()
    key = text.upper().replace('0', 'O').replace('-', ' ')
    key = ' '.join(key.split())
    if 'OPEN' in key:
        return 'Opening'
    if 'ONGOING' in key.replace(' ', ''):
        return 'Ongoing'
    if 'FINISH' in key:
        return 'Finished'
    if 'INSPECT' in key:
        return 'Inspected'
    return text[:20]


def _is_station_label(value):
    import re
    return bool(value and re.match(r'^\d+\+\d+$', str(value).strip()))


def _import_revetment(project, file_content, duplicate_action, user):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb['REVETMENT'] if 'REVETMENT' in wb.sheetnames else wb.active

    success = failed = 0
    errors = []
    activities_by_col = {}

    try:
        for col in range(3, 19):
            name = str(ws.cell(row=4, column=col).value or '').strip()
            if not name:
                continue
            group_name = str(_merged_value(ws, 3, col) or '').strip()
            template_column = get_column_letter(col)
            is_inspection = 'INSPECTION' in name.upper()
            activity, _ = RevetmentActivity.objects.update_or_create(
                project=project,
                template_column=template_column,
                defaults={
                    'group_name': group_name[:120],
                    'name': name[:120],
                    'unit': 'status' if is_inspection else '',
                    'is_inspection': is_inspection,
                    'sort_order': col,
                    'is_active': True,
                },
            )
            activities_by_col[col] = activity

        current_sop = ''
        for row in range(5, ws.max_row + 1):
            sop_candidate = str(ws.cell(row=row, column=3).value or '').strip()
            if sop_candidate.upper().startswith('SOP'):
                current_sop = sop_candidate[:50]

            station_value = ws.cell(row=row, column=2).value
            if not _is_station_label(station_value):
                continue
            station_label = str(station_value).strip()
            station, _ = RevetmentStation.objects.update_or_create(
                project=project,
                station=station_label,
                defaults={
                    'sop': current_sop,
                    'sort_order': row,
                    'is_active': True,
                },
            )
            success += 1

            for col, activity in activities_by_col.items():
                value = ws.cell(row=row, column=col).value
                if value is None or value == '':
                    continue
                status = _normalize_revetment_status(value)
                inspection_date = _to_date(value) if activity.is_inspection or _to_date(value) else None
                record_date = inspection_date or date.today()

                record, _ = RevetmentDailyRecord.objects.get_or_create(
                    project=project,
                    record_date=record_date,
                    defaults={
                        'day_name': record_date.strftime('%a'),
                        'remarks': 'Imported from Revetment workbook',
                        'created_by': user,
                    },
                )
                defaults = {
                    'status': status,
                    'inspection_date': inspection_date if activity.is_inspection else None,
                    'remarks': str(value)[:300],
                }
                if duplicate_action == 'update':
                    RevetmentDailyItem.objects.update_or_create(
                        record=record,
                        station=station,
                        activity=activity,
                        defaults=defaults,
                    )
                    success += 1
                else:
                    _, created = RevetmentDailyItem.objects.get_or_create(
                        record=record,
                        station=station,
                        activity=activity,
                        defaults=defaults,
                    )
                    success += 1 if created else 0
    except Exception as exc:
        failed += 1
        errors.append(str(exc)[:200])

    return success, failed, '; '.join(errors[:5])
