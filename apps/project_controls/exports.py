import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from .models import (
    RockDailyRecord, SandDailyRecord,
    RevetmentStation, RevetmentActivity, RevetmentDailyItem,
)


HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SUBHEADER_FILL = PatternFill('solid', start_color='2E75B6')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=9)
TOTAL_FILL = PatternFill('solid', start_color='D6E4F0')
TOTAL_FONT = Font(bold=True, size=9)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _apply_header(ws, row, col, value, fill=None, font=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell


def export_rock_excel(project):
    records = RockDailyRecord.objects.filter(project=project).order_by('record_date')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rock Summary'

    # Title
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = f'ROCK SUMMARY – {project.project_name} ({project.contract_no})'
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')

    headers = [
        'Day', 'Date',
        'Material Type', 'Source / Quarry', 'Destination / Placement Area',
        'TCT Daily (Ton)', 'TCT Accum (Ton)', 'Trips', 'Trucks',
        'Placed Daily (Ton)', 'Placed Accum (Ton)',
        'Station of Core',
        'Core Outside Daily', 'Core Outside Accum', 'Core Inside Accum',
        'Stock Balance', 'Remarks',
    ]
    for col, h in enumerate(headers, 1):
        _apply_header(ws, 2, col, h, HEADER_FILL, HEADER_FONT, Alignment(horizontal='center', wrap_text=True))
    ws.row_dimensions[2].height = 30

    for r, rec in enumerate(records, 3):
        data = [
            rec.day_name, str(rec.record_date),
            rec.material_type, rec.source_quarry, rec.destination_area,
            float(rec.tct_daily_ton), float(rec.tct_accum_ton), rec.tct_trips, rec.tct_trucks,
            float(rec.placed_daily_ton), float(rec.placed_accum_ton),
            rec.station_of_core,
            float(rec.core_outside_daily), float(rec.core_outside_accum), float(rec.core_inside_accum),
            float(rec.stock_balance), rec.remarks,
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Auto column widths
    col_widths = [8, 12, 16, 20, 24, 14, 14, 8, 8, 14, 14, 20, 14, 14, 14, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_sand_excel(project):
    records = SandDailyRecord.objects.filter(project=project).order_by('record_date')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sand Summary'

    ws.merge_cells('A1:V1')
    title_cell = ws['A1']
    title_cell.value = f'SAND SUMMARY – {project.project_name} ({project.contract_no})'
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')

    headers = [
        'Day', 'Date',
        'TCT Pier Daily', 'TCT Pier Accum', 'TCT Pier Trips', 'TCT Pier Trucks',
        'Stockpile Sand MTP3 Daily', 'Stockpile Sand MTP3 Accum', 'MTP3 Trips', 'MTP3 Trucks',
        'Chalothon Sand Pit Daily', 'Chalothon Sand Pit Accum',
        'Khlong Bang Phai (Oswald) Daily', 'Khlong Bang Phai (Oswald) Accum',
        'Total Daily', 'Total Accum',
        'Offshore Daily', 'Offshore Accum',
        'Onshore Daily', 'Onshore Accum',
        'Remaining TCT Pier', 'Remaining Stockpile Sand MTP3',
    ]
    for col, h in enumerate(headers, 1):
        _apply_header(ws, 2, col, h, HEADER_FILL, HEADER_FONT, Alignment(horizontal='center', wrap_text=True))
    ws.row_dimensions[2].height = 30

    for r, rec in enumerate(records, 3):
        data = [
            rec.day_name, str(rec.record_date),
            float(rec.tct_daily_ton), float(rec.tct_accum_ton), rec.tct_trips, rec.tct_trucks,
            float(rec.mtp3_daily_ton), float(rec.mtp3_accum_ton), rec.mtp3_trips, rec.mtp3_trucks,
            float(rec.chalothon_daily_ton), float(rec.chalothon_accum_ton),
            float(rec.khlong_bang_phai_daily_ton), float(rec.khlong_bang_phai_accum_ton),
            float(rec.total_daily_ton), float(rec.total_accum_ton),
            float(rec.offshore_daily_ton), float(rec.offshore_accum_ton),
            float(rec.onshore_daily_ton), float(rec.onshore_accum_ton),
            float(rec.remaining_tct), float(rec.remaining_mtp3),
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_rap_excel(plan):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Recovery Action Plan'

    ws.merge_cells('A1:I1')
    ws['A1'].value = f'RECOVERY ACTION PLAN – {plan.title}'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'].value = (f'Project: {plan.project.project_name}  |  '
                      f'Period: {plan.start_date} to {plan.end_date}  |  Status: {plan.status}')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Item', 'Description', 'Total Qty', 'Unit',
               'Total Planned', 'Total Actual', 'Remaining', '% Complete', 'Status']
    for col, h in enumerate(headers, 1):
        _apply_header(ws, 4, col, h, HEADER_FILL, HEADER_FONT, Alignment(horizontal='center'))

    for r, item in enumerate(plan.items.order_by('item_no'), 5):
        row_data = [
            item.item_no, item.description, float(item.total_quantity), item.unit,
            float(item.total_planned), float(item.total_actual),
            float(item.remaining_quantity), item.percent_complete, item.latest_status,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    widths = [6, 50, 12, 8, 12, 12, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_revetment_excel(project):
    stations = list(RevetmentStation.objects.filter(project=project, is_active=True))
    activities = list(RevetmentActivity.objects.filter(project=project, is_active=True))
    items = (
        RevetmentDailyItem.objects
        .filter(record__project=project)
        .select_related('record', 'station', 'activity')
        .order_by('record__record_date', 'created_at', 'pk')
    )

    quantity_map = {}
    latest_map = {}
    for item in items:
        key = (item.station_id, item.activity_id)
        if item.quantity_done is not None:
            quantity_map[key] = quantity_map.get(key, 0) + item.quantity_done
        latest_map[key] = item

    wb = Workbook()
    ws = wb.active
    ws.title = 'REVETMENT'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, len(activities) + 3))
    ws.cell(1, 1, f'REVETMENT PROGRESS - {project.project_name} ({project.contract_no})')
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    ws.cell(3, 1, 'DATE')
    ws.cell(3, 2, 'STA.')
    ws.cell(3, 3, 'SOP')
    for cell in ws[3]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    start_col = 4
    group_start = start_col
    current_group = None
    for idx, activity in enumerate(activities, start_col):
        if current_group is None:
            current_group = activity.group_name
            group_start = idx
        if activity.group_name != current_group:
            if group_start < idx - 1:
                ws.merge_cells(start_row=3, start_column=group_start, end_row=3, end_column=idx - 1)
            ws.cell(3, group_start, current_group or 'REVETMENT')
            current_group = activity.group_name
            group_start = idx
        ws.cell(4, idx, activity.name)
    if activities:
        if group_start < start_col + len(activities) - 1:
            ws.merge_cells(start_row=3, start_column=group_start, end_row=3, end_column=start_col + len(activities) - 1)
        ws.cell(3, group_start, current_group or 'REVETMENT')
    remark_col = start_col + len(activities)
    ws.cell(3, remark_col, 'REMARK')

    for col in range(start_col, remark_col + 1):
        ws.cell(3, col).font = HEADER_FONT
        ws.cell(3, col).fill = HEADER_FILL
        ws.cell(3, col).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.cell(3, col).border = THIN_BORDER
        ws.cell(4, col).font = SUBHEADER_FONT if col < remark_col else HEADER_FONT
        ws.cell(4, col).fill = SUBHEADER_FILL if col < remark_col else HEADER_FILL
        ws.cell(4, col).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.cell(4, col).border = THIN_BORDER

    for row_no, station in enumerate(stations, 5):
        latest_station_item = None
        for activity in activities:
            candidate = latest_map.get((station.pk, activity.pk))
            if candidate:
                latest_station_item = candidate
        ws.cell(row_no, 1, latest_station_item.record.record_date if latest_station_item else None)
        ws.cell(row_no, 2, station.station)
        ws.cell(row_no, 3, station.sop)
        for idx, activity in enumerate(activities, start_col):
            latest = latest_map.get((station.pk, activity.pk))
            total = quantity_map.get((station.pk, activity.pk))
            if activity.is_inspection:
                value = latest.inspection_date if latest and latest.inspection_date else (latest.status if latest else None)
            else:
                value = float(total) if total else (latest.status if latest else None)
            cell = ws.cell(row_no, idx, value)
            if isinstance(value, float):
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal='right')
            cell.border = THIN_BORDER
        ws.cell(row_no, remark_col, latest_station_item.remarks if latest_station_item else '')
        for col in range(1, remark_col + 1):
            ws.cell(row_no, col).border = THIN_BORDER

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    for col in range(start_col, remark_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    raw = wb.create_sheet('Daily Entries')
    headers = ['Date', 'Project', 'SOP', 'Station', 'Group', 'Activity', 'Qty',
               'Unit', 'Status', 'Inspection Date', 'Remarks']
    for col, header in enumerate(headers, 1):
        _apply_header(raw, 1, col, header, HEADER_FILL, HEADER_FONT, Alignment(horizontal='center'))
    for row_no, item in enumerate(items, 2):
        values = [
            item.record.record_date,
            project.project_name,
            item.station.sop,
            item.station.station,
            item.activity.group_name,
            item.activity.name,
            float(item.quantity_done) if item.quantity_done is not None else None,
            item.activity.display_unit,
            item.status,
            item.inspection_date,
            item.remarks,
        ]
        for col, value in enumerate(values, 1):
            cell = raw.cell(row_no, col, value)
            cell.border = THIN_BORDER
            if col == 7 and value is not None:
                cell.number_format = '#,##0.000'
    for col in range(1, len(headers) + 1):
        raw.column_dimensions[get_column_letter(col)].width = 16
    raw.column_dimensions['F'].width = 24
    raw.column_dimensions['K'].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_rock_pdf(record):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(
        f'<b>ROCK DAILY SUMMARY</b> – {record.project.project_name}',
        styles['Title']
    ))
    story.append(Paragraph(f'Date: {record.record_date}   |   Status: Report', styles['Normal']))
    story.append(Spacer(1, 0.3*cm))

    summary_data = [
        ['Field', 'Value'],
        ['TCT Daily Delivered (Ton)', f'{record.tct_daily_ton:,.2f}'],
        ['TCT Accumulative (Ton)', f'{record.tct_accum_ton:,.2f}'],
        ['Trips / Trucks', f'{record.tct_trips} / {record.tct_trucks}'],
        ['Rock Placed Daily (Ton)', f'{record.placed_daily_ton:,.2f}'],
        ['Rock Placed Accumulative (Ton)', f'{record.placed_accum_ton:,.2f}'],
        ['Stock Balance (Ton)', f'{record.stock_balance:,.2f}'],
        ['Station of Core', record.station_of_core or '-'],
        ['Core Outside Plot Tank (Accum)', f'{record.core_outside_accum:,.2f}'],
        ['Core Inside Plot Tank (Accum)', f'{record.core_inside_accum:,.2f}'],
    ]
    t = Table(summary_data, colWidths=[8*cm, 6*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    placements = record.barge_placements.select_related('barge').all()
    if placements:
        story.append(Paragraph('<b>Barge Placements</b>', styles['Heading3']))
        bp_data = [['Barge', 'Quantity (Ton)', 'Trips', 'Station']]
        for bp in placements:
            bp_data.append([bp.barge.name, f'{bp.quantity_ton:,.2f}', str(bp.trips), bp.station or '-'])
        bt = Table(bp_data, colWidths=[5*cm, 4*cm, 3*cm, 8*cm])
        bt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(bt)

    if record.remarks:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f'<b>Remarks:</b> {record.remarks}', styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return buf


def export_rap_pdf(plan):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f'<b>RECOVERY ACTION PLAN</b>', styles['Title']))
    story.append(Paragraph(f'{plan.title}', styles['Heading2']))
    story.append(Paragraph(
        f'Project: {plan.project.project_name}  |  '
        f'Period: {plan.start_date} to {plan.end_date}  |  Status: {plan.status}',
        styles['Normal']
    ))
    if plan.prepared_by:
        story.append(Paragraph(f'Prepared by: {plan.prepared_by.get_full_name()}', styles['Normal']))
    if plan.approved_by:
        story.append(Paragraph(f'Approved by: {plan.approved_by.get_full_name()}', styles['Normal']))
    story.append(Spacer(1, 0.4*cm))

    table_data = [['Item', 'Description', 'Total Qty', 'Unit',
                   'Planned', 'Actual', 'Remaining', '% Done', 'Status']]
    for item in plan.items.order_by('item_no'):
        table_data.append([
            item.item_no,
            Paragraph(item.description, styles['Normal']),
            f'{item.total_quantity:,.0f}',
            item.unit,
            f'{item.total_planned:,.0f}',
            f'{item.total_actual:,.0f}',
            f'{item.remaining_quantity:,.0f}',
            f'{item.percent_complete:.1f}%',
            item.latest_status,
        ])

    col_widths = [1.5*cm, 8*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm]
    t = Table(table_data, colWidths=col_widths)
    status_colors = {'Ahead': colors.green, 'On Track': colors.blue,
                     'Delayed': colors.red, 'No Activity': colors.grey}
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf
