import io
import os

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)

_REGISTERED_FONTS = False


def _register_fonts():
    global _REGISTERED_FONTS
    if _REGISTERED_FONTS:
        return

    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    fonts_dir = os.path.join(root_dir, 'static', 'fonts')
    regular = os.path.join(fonts_dir, 'THSarabunNew.ttf')
    bold = os.path.join(fonts_dir, 'THSarabunNew-Bold.ttf')
    if os.path.exists(regular):
        pdfmetrics.registerFont(TTFont('THSarabunNew', regular))
        pdfmetrics.registerFont(TTFont('THSarabunNew-Bold', bold if os.path.exists(bold) else regular))
    _REGISTERED_FONTS = True


def _pdf_styles():
    _register_fonts()
    font = 'THSarabunNew' if 'THSarabunNew' in pdfmetrics.getRegisteredFontNames() else 'Helvetica'
    bold = 'THSarabunNew-Bold' if 'THSarabunNew-Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold'
    return {
        'title': ParagraphStyle('MaterialTitle', fontName=bold, fontSize=18, leading=22, alignment=TA_CENTER,
                                textColor=colors.HexColor('#1F4E79'), spaceAfter=4),
        'meta': ParagraphStyle('MaterialMeta', fontName=font, fontSize=11, leading=14, alignment=TA_CENTER,
                               textColor=colors.HexColor('#4B5563')),
        'header': ParagraphStyle('MaterialHeader', fontName=bold, fontSize=10, leading=12, alignment=TA_CENTER,
                                 textColor=colors.white),
        'cell': ParagraphStyle('MaterialCell', fontName=font, fontSize=9, leading=11),
        'cell_right': ParagraphStyle('MaterialCellRight', fontName=font, fontSize=9, leading=11,
                                     alignment=TA_RIGHT),
    }, font, bold


def _p(value, style):
    text = '' if value is None else str(value)
    safe = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return Paragraph(safe, style)


def _delivery_rows(deliveries):
    for delivery in deliveries:
        source_parts = []
        if delivery.source_area:
            source_parts.append(delivery.source_area.name)
        if delivery.source:
            source_parts.append(delivery.source)
        yield {
            'date': delivery.delivery_date,
            'time': delivery.delivery_time.strftime('%H:%M') if delivery.delivery_time else '',
            'project': delivery.project.contract_no,
            'project_name': delivery.project.project_name,
            'material': delivery.material.name,
            'source': ' / '.join(source_parts),
            'dn': delivery.delivery_note_no,
            'truck': delivery.truck_no,
            'quantity': delivery.quantity,
            'unit': delivery.material.unit,
            'unit_price': delivery.unit_price,
            'amount': delivery.total_amount,
            'remarks': delivery.remarks,
        }


def export_material_deliveries_excel(deliveries):
    rows = list(_delivery_rows(deliveries))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Material Deliveries'

    ws.merge_cells('A1:M1')
    ws['A1'] = 'รายการรับวัสดุ'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = f'Export: {timezone.localtime().strftime("%Y-%m-%d %H:%M")} | Records: {len(rows):,}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = [
        'วันที่', 'เวลา', 'โครงการ', 'ชื่อโครงการ', 'วัสดุ', 'แหล่งวัสดุ',
        'เลขที่ใบส่งของ', 'ทะเบียนรถ', 'ปริมาณ', 'หน่วย', 'ราคา/หน่วย', 'มูลค่า', 'หมายเหตุ',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_no, row in enumerate(rows, 5):
        values = [
            row['date'], row['time'], row['project'], row['project_name'], row['material'], row['source'],
            row['dn'], row['truck'], row['quantity'], row['unit'], row['unit_price'], row['amount'], row['remarks'],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_no, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=col in {4, 6, 13})
            if col == 1:
                cell.number_format = 'yyyy-mm-dd'
            if col in {9, 11, 12}:
                cell.number_format = '#,##0.000' if col == 9 else '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='top')

    widths = [12, 8, 14, 28, 18, 26, 18, 14, 14, 10, 14, 14, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:M{max(4, len(rows) + 4)}'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_material_deliveries_pdf(deliveries):
    rows = list(_delivery_rows(deliveries))
    styles, font, bold = _pdf_styles()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )

    story = [
        Paragraph('รายการรับวัสดุ', styles['title']),
        Paragraph(
            f'Export: {timezone.localtime().strftime("%Y-%m-%d %H:%M")} | Records: {len(rows):,} | Page size: A4',
            styles['meta'],
        ),
        Spacer(1, 0.25 * cm),
    ]

    headers = ['วันที่', 'โครงการ', 'วัสดุ', 'แหล่งวัสดุ', 'ใบส่งของ', 'ทะเบียนรถ', 'ปริมาณ', 'ราคา/หน่วย', 'มูลค่า']
    table_data = [[_p(header, styles['header']) for header in headers]]
    for row in rows:
        qty = f"{row['quantity']:,.3f} {row['unit']}"
        unit_price = f"{row['unit_price']:,.2f}" if row['unit_price'] is not None else '-'
        amount = f"{row['amount']:,.2f}" if row['amount'] is not None else '-'
        date_text = row['date'].strftime('%Y-%m-%d')
        if row['time']:
            date_text = f"{date_text}\n{row['time']}"
        table_data.append([
            _p(date_text, styles['cell']),
            _p(row['project'], styles['cell']),
            _p(row['material'], styles['cell']),
            _p(row['source'] or '-', styles['cell']),
            _p(row['dn'] or '-', styles['cell']),
            _p(row['truck'] or '-', styles['cell']),
            _p(qty, styles['cell_right']),
            _p(unit_price, styles['cell_right']),
            _p(amount, styles['cell_right']),
        ])

    table = Table(
        table_data,
        colWidths=[2.1 * cm, 2.1 * cm, 3.1 * cm, 4.6 * cm, 2.7 * cm, 2.4 * cm, 2.9 * cm, 2.6 * cm, 2.6 * cm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), bold),
        ('FONTNAME', (0, 1), (-1, -1), font),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D1D5DB')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf
