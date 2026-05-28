"""
Excel import helper for Contractor Daily Reports.
Parses the official template workbook and returns structured data.
"""
import openpyxl
from decimal import Decimal, InvalidOperation
from datetime import date, datetime as _datetime


WEATHER_MAP = {v.upper(): v for v in ['Clear', 'Cloudy', 'Windy', 'Raining', 'High Wave', 'Other']}


def _cell_str(sheet, row, col):
    val = sheet.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ''


def _to_decimal(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def parse_import_file(uploaded_file):
    """
    Parse an uploaded Excel file using the official contractor daily report template.

    Returns a dict with keys:
        report_date, contractor_name, weather_day, weather_night, wind_speed,
        total_manpower, prepared_by_name, checked_by_name, remarks,
        equipment_items, manpower_items, activities, lookaheads

    Raises ValueError with a human-readable message if the file is invalid.
    """
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read Excel file: {exc}")

    required_sheets = {'Report Info', 'Work Activities'}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        raise ValueError(
            f"Sheet(s) not found: {', '.join(missing)}. "
            "Please use the official import template."
        )

    # ── Sheet 1: Report Info ──────────────────────────────────────────────────
    ws = wb['Report Info']

    raw_date = ws.cell(3, 2).value
    # openpyxl returns date cells as datetime objects (datetime subclasses date),
    # so check datetime FIRST to avoid getting the time component in isoformat.
    if isinstance(raw_date, _datetime):
        report_date = raw_date.date()
    elif isinstance(raw_date, date):
        report_date = raw_date
    elif raw_date:
        try:
            s = str(raw_date).strip()
            # Handle strings that include a time component (e.g. "2026-05-11T00:00:00")
            if 'T' in s or (' ' in s and len(s) > 10):
                report_date = _datetime.fromisoformat(s).date()
            else:
                report_date = date.fromisoformat(s)
        except ValueError:
            raise ValueError(
                f"Invalid report date '{raw_date}'. "
                "Use YYYY-MM-DD format in the 'Report Date' cell."
            )
    else:
        raise ValueError("Report Date is required (row 3 of 'Report Info' sheet).")

    contractor_name = _cell_str(ws, 4, 2)
    weather_day = WEATHER_MAP.get(_cell_str(ws, 5, 2).upper(), 'Clear')
    weather_night = WEATHER_MAP.get(_cell_str(ws, 6, 2).upper(), 'Clear')
    wind_speed = _cell_str(ws, 7, 2)
    total_manpower = _to_int(ws.cell(8, 2).value)
    prepared_by_name = _cell_str(ws, 9, 2)
    checked_by_name = _cell_str(ws, 10, 2)
    remarks = _cell_str(ws, 11, 2)

    # ── Sheet 2: Equipment ────────────────────────────────────────────────────
    equipment_items = []
    if 'Equipment' in wb.sheetnames:
        for row in wb['Equipment'].iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name:
                continue
            equipment_items.append({
                'equipment_name': str(name).strip(),
                'quantity': _to_int(row[1] if len(row) > 1 else None, 1),
                'remarks': str(row[2]).strip() if len(row) > 2 and row[2] else '',
            })

    # ── Sheet 3: Manpower ─────────────────────────────────────────────────────
    manpower_items = []
    if 'Manpower' in wb.sheetnames:
        for row in wb['Manpower'].iter_rows(min_row=2, values_only=True):
            role = row[0]
            if not role:
                continue
            manpower_items.append({
                'role': str(role).strip(),
                'quantity': _to_int(row[1] if len(row) > 1 else None),
                'company': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'remarks': str(row[3]).strip() if len(row) > 3 and row[3] else '',
            })

    # ── Sheet 4: Work Activities ──────────────────────────────────────────────
    activities = []
    for i, row in enumerate(wb['Work Activities'].iter_rows(min_row=2, values_only=True), start=1):
        desc = row[1] if len(row) > 1 else row[0]
        if not desc:
            continue
        activities.append({
            'item_no': _to_int(row[0], i) if row[0] else i,
            'description': str(desc).strip(),
            'location': str(row[2]).strip() if len(row) > 2 and row[2] else '',
            'quantity': _to_decimal(row[3] if len(row) > 3 else None),
            'unit': str(row[4]).strip() if len(row) > 4 and row[4] else '',
            'problem': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            'remarks': str(row[6]).strip() if len(row) > 6 and row[6] else '',
        })

    # ── Sheet 5: Lookahead ────────────────────────────────────────────────────
    lookaheads = []
    if 'Lookahead' in wb.sheetnames:
        for i, row in enumerate(wb['Lookahead'].iter_rows(min_row=2, values_only=True), start=1):
            desc = row[1] if len(row) > 1 else row[0]
            if not desc:
                continue
            lookaheads.append({
                'item_no': _to_int(row[0], i) if row[0] else i,
                'description': str(desc).strip(),
                'location': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'quantity': _to_decimal(row[3] if len(row) > 3 else None),
                'unit': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                'remarks': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            })

    return {
        'report_date': report_date,
        'contractor_name': contractor_name,
        'weather_day': weather_day,
        'weather_night': weather_night,
        'wind_speed': wind_speed,
        'total_manpower': total_manpower,
        'prepared_by_name': prepared_by_name,
        'checked_by_name': checked_by_name,
        'remarks': remarks,
        'equipment_items': equipment_items,
        'manpower_items': manpower_items,
        'activities': activities,
        'lookaheads': lookaheads,
    }


def build_template_workbook():
    """Return an openpyxl Workbook containing the blank import template."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1E40AF')
    lbl_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=13, color='1E3A8A')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFDBFE')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, value):
        cell.value = value
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = bdr

    def lbl(cell, value):
        cell.value = value
        cell.font = lbl_font
        cell.alignment = left
        cell.border = bdr

    def data_cell(cell, value=''):
        cell.value = value
        cell.alignment = left
        cell.border = bdr

    # ── Sheet 1: Report Info ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Report Info'
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 42
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A1:B1')
    ws['A1'] = 'CONTRACTOR DAILY REPORT – IMPORT TEMPLATE'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:B2')
    ws['A2'] = (
        'Fill in column B for each row, then complete the other sheets. '
        'Do not change sheet names or row order.'
    )
    ws['A2'].font = Font(italic=True, size=9, color='6B7280')
    ws['A2'].alignment = left
    ws.row_dimensions[2].height = 14

    rows = [
        ('Report Date (YYYY-MM-DD)', ''),
        ('Contractor Name', ''),
        ('Weather – Day', 'CLEAR / CLOUDY / WINDY / RAINING / HIGH WAVE'),
        ('Weather – Night', 'CLEAR / CLOUDY / WINDY / RAINING / HIGH WAVE'),
        ('Wind Speed (km/hr)', ''),
        ('Total Manpower (persons)', ''),
        ('Prepared By', ''),
        ('Checked By', ''),
        ('Remarks', ''),
    ]
    for r, (label, hint) in enumerate(rows, start=3):
        lbl(ws.cell(r, 1), label)
        data_cell(ws.cell(r, 2), hint)
        ws.row_dimensions[r].height = 18

    # ── Sheet 2: Equipment ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Equipment')
    for col, w in zip('ABCD', [35, 12, 30, 1]):
        ws2.column_dimensions[col].width = w
    for c, h in enumerate(['Equipment Name', 'Quantity', 'Remarks'], 1):
        hdr(ws2.cell(1, c), h)
    ws2.row_dimensions[1].height = 20
    for r, (name, qty, rem) in enumerate([
        ('Excavator', 15, ''),
        ('Dump Truck (10-wheel)', 10, ''),
        ('Mobile Crane', 1, ''),
        ('Bulldozer D6', 1, ''),
        ('Water Truck', 4, ''),
    ], start=2):
        for c, v in enumerate([name, qty, rem], 1):
            data_cell(ws2.cell(r, c), v)

    # ── Sheet 3: Manpower ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Manpower')
    for col, w in zip('ABCD', [30, 12, 28, 28]):
        ws3.column_dimensions[col].width = w
    for c, h in enumerate(['Role / Position', 'Quantity', 'Company', 'Remarks'], 1):
        hdr(ws3.cell(1, c), h)
    ws3.row_dimensions[1].height = 20
    for r, row in enumerate([
        ('Project Manager', 1, '', ''),
        ('Construction Manager', 1, '', ''),
        ('Engineer', 1, '', ''),
        ('SHE Manager', 1, '', ''),
        ('SHE Officer', 4, '', ''),
        ('Operator', 15, '', ''),
        ('Driver', 8, '', ''),
        ('Worker', 57, '', ''),
        ('Marine Worker', 26, '', ''),
    ], start=2):
        for c, v in enumerate(row, 1):
            data_cell(ws3.cell(r, c), v)

    # ── Sheet 4: Work Activities ──────────────────────────────────────────────
    ws4 = wb.create_sheet('Work Activities')
    for col, w in zip('ABCDEFG', [6, 48, 22, 14, 10, 28, 28]):
        ws4.column_dimensions[col].width = w
    for c, h in enumerate(['No.', 'Description', 'Location', 'Quantity', 'Unit', 'Problem', 'Remarks'], 1):
        hdr(ws4.cell(1, c), h)
    ws4.row_dimensions[1].height = 20
    for r, row in enumerate([
        (1, 'Rock transport from quarry', 'Port TCT', 1836.35, 'ton', '', '10 trucks, 53 trips'),
        (2, 'Rock loading onto barge – MAP TA PHUT 8', '', 650.00, 'ton', '', ''),
        (3, 'Rock loading onto barge – MAP TA PHUT 10', '', 600.00, 'ton', '', ''),
        (4, 'Bedding + Core rock placement (seabed)', 'TTT', 1800.00, 'ton', '', ''),
        (5, 'Sand transport – land route', 'SOP01-SOP04', 1690.50, 'ton', '', '11 trucks, 69 trips'),
    ], start=2):
        for c, v in enumerate(row, 1):
            data_cell(ws4.cell(r, c), v)

    # ── Sheet 5: Lookahead ────────────────────────────────────────────────────
    ws5 = wb.create_sheet('Lookahead')
    for col, w in zip('ABCDEF', [6, 48, 22, 14, 10, 28]):
        ws5.column_dimensions[col].width = w
    for c, h in enumerate(['No.', 'Work Activity', 'Location', 'Quantity', 'Unit', 'Remarks'], 1):
        hdr(ws5.cell(1, c), h)
    ws5.row_dimensions[1].height = 20
    for r, row in enumerate([
        (1, 'Rock transport from quarry (Thawarafn)', '', 1100, 'ton', ''),
        (2, 'Rock loading onto barge J.YUTTACHAI', '', 700, 'ton', ''),
        (3, 'Bedding + Core rock placement', 'TTT seabed', 1300, 'ton', ''),
        (4, 'Sand production from Chalotra pit', '', 6500, 'ton', ''),
        (5, 'Sand transport – land route', 'SOP01-SOP04', 5570, 'ton', ''),
    ], start=2):
        for c, v in enumerate(row, 1):
            data_cell(ws5.cell(r, c), v)

    return wb
