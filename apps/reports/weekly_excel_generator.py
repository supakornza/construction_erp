"""
Weekly Site Report Excel generator.

Sheets:
  1. Summary      – project info, weekly KPIs
  2. Activities   – daily work activities for the week
  3. Manpower     – daily manpower by category, weekly totals
  4. Equipment    – daily equipment status/hours, weekly totals
  5. Safety       – toolbox meetings, inspections, incidents

Entry point: generate_weekly_excel(project, week_start) -> io.BytesIO
week_start must be a Monday (datetime.date).
"""
import io
from datetime import timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
_NAVY       = '1B2A4A'
_NAVY_LIGHT = 'DCE8F5'
_GREEN      = '1E8449'
_GREEN_LIGHT= 'D4EDDA'
_YELLOW     = 'FFF3CD'
_RED_LIGHT  = 'F8D7DA'
_GREY_LIGHT = 'F5F7FA'
_GREY_TEXT  = '6C757D'
_WHITE      = 'FFFFFF'
_ORANGE     = 'E67E22'

# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=10, color='000000', italic=False):
    return Font(name='Arial', size=size, bold=bold, color=color, italic=italic)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border():
    s = Side(style='thin', color='B0BEC5')
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, values, widths=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=True, size=9, color=_WHITE)
        c.fill      = _fill(_NAVY)
        c.alignment = _align('center', 'center', wrap=True)
        c.border    = _border()
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _cell(ws, row, col, value='', bold=False, fill=None, align='left', wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=9)
    c.alignment = _align(align, 'center', wrap)
    c.border    = _border()
    if fill:
        c.fill = _fill(fill)
    return c


def _section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=10, color=_WHITE)
    c.fill      = _fill(_NAVY)
    c.alignment = _align('left', 'center')
    ws.row_dimensions[row].height = 16


def _week_dates(week_start):
    return [week_start + timedelta(days=i) for i in range(7)]


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _build_summary(wb, project, week_start, reports, week_dates):
    from apps.manpower.models import DailyManpowerRecord
    from apps.equipment.models import DailyEquipmentRecord
    from apps.safety.models import ToolboxMeeting, IncidentReport
    from django.db.models import Sum

    ws = wb.active
    ws.title = 'Weekly Summary'
    ws.sheet_properties.tabColor = _NAVY

    week_end = week_dates[-1]
    NCOLS = 6

    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage  = True

    col_ws = [20, 20, 15, 15, 15, 20]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value='WEEKLY SITE REPORT')
    c.font = _font(bold=True, size=16, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 26
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=project.project_name)
    c.font = _font(bold=True, size=12, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 18
    r += 2

    # Project info
    _section(ws, r, '1.  PROJECT INFORMATION', NCOLS); r += 1
    pairs = [
        ('Contract No',   project.contract_no,   'Owner',       project.owner),
        ('Contractor',    project.contractor,     'Consultant',  project.consultant or '—'),
        ('Week Start',    str(week_start),        'Week End',    str(week_end)),
        ('Report Period', f'{week_start} to {week_end}', 'Project Status', project.status),
    ]
    for lbl1, val1, lbl2, val2 in pairs:
        for col, (txt, is_lbl) in enumerate(
            [(lbl1, True), (val1, False), (lbl2, True), (val2, False)], 1
        ):
            real_col = [1, 2, 4, 5][col - 1]
            c = ws.cell(row=r, column=real_col, value=txt)
            c.font      = _font(bold=is_lbl, size=9, color=_NAVY if is_lbl else '000000')
            c.fill      = _fill(_NAVY_LIGHT) if is_lbl else _fill(_WHITE)
            c.alignment = _align('left', 'center')
            c.border    = _border()
        r += 1
    r += 1

    # Daily report coverage
    _section(ws, r, '2.  DAILY REPORT COVERAGE', NCOLS); r += 1
    _hdr(ws, r, ['Date', 'Day', 'Report Status', 'Weather AM', 'Weather PM', 'Remarks']); r += 1
    for d in week_dates:
        rep = next((rp for rp in reports if rp.report_date == d), None)
        day_name = d.strftime('%A')
        if rep:
            bg = _GREEN_LIGHT if rep.status == 'Approved' else _YELLOW if rep.status == 'Submitted' else None
            vals = [str(d), day_name, rep.status, rep.weather_morning, rep.weather_afternoon, rep.remarks[:100] if rep.remarks else '']
        else:
            bg = _GREY_LIGHT
            vals = [str(d), day_name, 'No Report', '—', '—', '']
        for col, val in enumerate(vals, 1):
            _cell(ws, r, col, val, fill=bg, align='center' if col <= 3 else 'left')
        r += 1
    r += 1

    # Weekly KPIs
    _section(ws, r, '3.  WEEKLY KEY FIGURES', NCOLS); r += 1
    total_mp = DailyManpowerRecord.objects.filter(
        project=project, report_date__range=(week_start, week_end)
    ).aggregate(total=Sum('quantity'))['total'] or 0

    total_eq_hrs = DailyEquipmentRecord.objects.filter(
        project=project, report_date__range=(week_start, week_end),
        status='Working',
    ).aggregate(total=Sum('working_hours'))['total'] or Decimal('0')

    toolbox_count = ToolboxMeeting.objects.filter(
        project=project, meeting_date__range=(week_start, week_end)
    ).count()

    incident_count = IncidentReport.objects.filter(
        project=project, incident_date__range=(week_start, week_end)
    ).count()

    kpi_rows = [
        ('Reports Filed',       len(reports),           f'out of 7 days'),
        ('Total Manpower',      int(total_mp),           'person-days'),
        ('Total Equipment Hrs', float(total_eq_hrs),     'working hours'),
        ('Toolbox Meetings',    toolbox_count,           'sessions'),
        ('Incidents',           incident_count,          'reported'),
    ]
    for kpi, val, unit in kpi_rows:
        _cell(ws, r, 1, kpi, bold=True, fill=_NAVY_LIGHT)
        _cell(ws, r, 2, val, bold=True, align='center')
        _cell(ws, r, 3, unit, fill=_GREY_LIGHT)
        for col in range(4, NCOLS + 1):
            _cell(ws, r, col, '')
        r += 1


# ── Sheet 2: Activities ───────────────────────────────────────────────────────

def _build_activities(wb, project, reports, week_dates):
    ws = wb.create_sheet(title='Daily Activities')
    ws.sheet_properties.tabColor = _GREEN

    NCOLS = 8
    col_ws = [12, 8, 35, 20, 8, 10, 10, 20]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value='WEEKLY WORK ACTIVITIES')
    c.font = _font(bold=True, size=13, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 22
    r += 2

    for d in week_dates:
        rep = next((rp for rp in reports if rp.report_date == d), None)
        day_label = f"{d.strftime('%A')}  {d}"

        _section(ws, r, day_label, NCOLS); r += 1
        if not rep:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
            c = ws.cell(row=r, column=1, value='No daily report recorded for this day.')
            c.font = _font(italic=True, size=9, color=_GREY_TEXT)
            c.fill = _fill(_GREY_LIGHT)
            c.border = _border()
            r += 2
            continue

        _hdr(ws, r, ['#', 'Date', 'Activity Description', 'Location', 'Unit', 'Quantity', '% Done', 'Remarks']); r += 1
        activities = rep.activities.select_related('work_area').all()
        if not activities:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
            _cell(ws, r, 1, 'No activities recorded.', fill=_GREY_LIGHT)
            r += 1
        else:
            for i, act in enumerate(activities, 1):
                loc = str(act.work_area) if act.work_area else '—'
                qty = float(act.quantity) if act.quantity is not None else None
                pct = float(act.percent_complete) if act.percent_complete is not None else None
                bg = _GREEN_LIGHT if pct and pct >= 100 else _YELLOW if pct and pct > 0 else None
                vals = [i, str(d), act.description, loc, act.unit or '—', qty, pct, act.remarks]
                for col, val in enumerate(vals, 1):
                    _cell(ws, r, col, val, fill=bg if col >= 6 else None,
                          align='center' if col in (1, 5, 6, 7) else 'left')
                r += 1
        r += 1


# ── Sheet 3: Manpower ─────────────────────────────────────────────────────────

def _build_manpower(wb, project, week_start, week_dates):
    from apps.manpower.models import DailyManpowerRecord
    from django.db.models import Sum

    ws = wb.create_sheet(title='Manpower')
    ws.sheet_properties.tabColor = '2874A6'

    # Collect categories used this week
    records = DailyManpowerRecord.objects.filter(
        project=project,
        report_date__range=(week_start, week_dates[-1]),
    ).select_related('category').order_by('report_date', 'category__name')

    categories = list({r.category for r in records})
    categories.sort(key=lambda c: c.name)

    day_labels = [d.strftime('%a\n%d/%m') for d in week_dates]
    headers = ['Category', 'Company'] + day_labels + ['Weekly Total']
    col_ws = [18, 20] + [9] * 7 + [12]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    c = ws.cell(row=r, column=1, value='WEEKLY MANPOWER SUMMARY')
    c.font = _font(bold=True, size=13, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 22
    r += 2

    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _hdr(ws, r, headers); r += 1

    # Group by category + company
    combos = {}
    for rec in records:
        key = (rec.category.get_name_display(), rec.company)
        if key not in combos:
            combos[key] = {d: 0 for d in week_dates}
        combos[key][rec.report_date] += rec.quantity

    daily_totals = {d: 0 for d in week_dates}
    for (cat, company), daily in combos.items():
        weekly = sum(daily.values())
        vals = [cat, company] + [daily[d] or '' for d in week_dates] + [weekly]
        for col, val in enumerate(vals, 1):
            _cell(ws, r, col, val, align='center' if col > 2 else 'left')
        for d in week_dates:
            daily_totals[d] += daily[d]
        r += 1

    # Total row
    grand_total = sum(daily_totals.values())
    tot_vals = ['TOTAL', ''] + [daily_totals[d] for d in week_dates] + [grand_total]
    for col, val in enumerate(tot_vals, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(_NAVY_LIGHT)
        c.alignment = _align('center' if col > 2 else 'left', 'center')
        c.border    = _border()


# ── Sheet 4: Equipment ────────────────────────────────────────────────────────

def _build_equipment(wb, project, week_start, week_dates):
    from apps.equipment.models import DailyEquipmentRecord
    from django.db.models import Sum

    ws = wb.create_sheet(title='Equipment')
    ws.sheet_properties.tabColor = _ORANGE

    records = DailyEquipmentRecord.objects.filter(
        project=project,
        report_date__range=(week_start, week_dates[-1]),
    ).select_related('equipment').order_by('equipment__name', 'report_date')

    day_labels = [d.strftime('%a\n%d/%m') for d in week_dates]
    headers = ['Equipment', 'Status (latest)'] + day_labels + ['Total Hrs']
    col_ws = [25, 15] + [9] * 7 + [10]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    c = ws.cell(row=r, column=1, value='WEEKLY EQUIPMENT SUMMARY')
    c.font = _font(bold=True, size=13, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 22
    r += 2

    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _hdr(ws, r, headers); r += 1

    # Group by equipment
    equip_map = {}
    for rec in records:
        eid = rec.equipment_id
        if eid not in equip_map:
            equip_map[eid] = {
                'name': str(rec.equipment),
                'status': rec.status,
                'daily': {d: Decimal('0') for d in week_dates},
            }
        equip_map[eid]['daily'][rec.report_date] += rec.working_hours or Decimal('0')
        equip_map[eid]['status'] = rec.status  # keep latest

    daily_totals = {d: Decimal('0') for d in week_dates}
    for eid, data in equip_map.items():
        total_hrs = sum(data['daily'].values())
        status_bg = _YELLOW if data['status'] in ('Standby', 'Breakdown') else None
        vals = [data['name'], data['status']] + [
            float(data['daily'][d]) if data['daily'][d] else ''
            for d in week_dates
        ] + [float(total_hrs)]
        for col, val in enumerate(vals, 1):
            _cell(ws, r, col, val,
                  fill=status_bg if col == 2 else None,
                  align='center' if col > 2 else 'left')
        for d in week_dates:
            daily_totals[d] += data['daily'][d]
        r += 1

    grand_total = sum(daily_totals.values())
    tot_vals = ['TOTAL', ''] + [float(daily_totals[d]) if daily_totals[d] else '' for d in week_dates] + [float(grand_total)]
    for col, val in enumerate(tot_vals, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(_NAVY_LIGHT)
        c.alignment = _align('center' if col > 2 else 'left', 'center')
        c.border    = _border()


# ── Sheet 5: Safety ───────────────────────────────────────────────────────────

def _build_safety(wb, project, week_start, week_dates):
    from apps.safety.models import ToolboxMeeting, SafetyInspection, IncidentReport

    ws = wb.create_sheet(title='Safety')
    ws.sheet_properties.tabColor = '922B21'

    week_end = week_dates[-1]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value='WEEKLY SAFETY SUMMARY')
    c.font = _font(bold=True, size=13, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 22
    r += 2

    col_ws = [12, 30, 20, 15, 15, 25]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Toolbox Meetings
    _section(ws, r, 'TOOLBOX MEETINGS', 6); r += 1
    _hdr(ws, r, ['Date', 'Topic', 'Location', 'Conducted By', 'Attendees', 'Remarks']); r += 1
    meetings = ToolboxMeeting.objects.filter(
        project=project, meeting_date__range=(week_start, week_end)
    ).select_related('conducted_by').order_by('meeting_date')
    if not meetings:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _cell(ws, r, 1, 'No toolbox meetings this week.', fill=_GREY_LIGHT, italic=True)
        r += 1
    else:
        for m in meetings:
            conductor = m.conducted_by.get_full_name() if m.conducted_by else '—'
            vals = [str(m.meeting_date), m.topic, m.location, conductor, m.attendee_count, m.remarks[:100]]
            for col, val in enumerate(vals, 1):
                _cell(ws, r, col, val, align='center' if col in (1, 5) else 'left')
            r += 1
    r += 1

    # Safety Inspections
    _section(ws, r, 'SAFETY INSPECTIONS', 6); r += 1
    _hdr(ws, r, ['Date', 'Finding', 'Category', 'Location', 'Status', 'Due Date']); r += 1
    inspections = SafetyInspection.objects.filter(
        project=project, inspection_date__range=(week_start, week_end)
    ).order_by('inspection_date')
    if not inspections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _cell(ws, r, 1, 'No safety inspections this week.', fill=_GREY_LIGHT)
        r += 1
    else:
        for insp in inspections:
            bg = _RED_LIGHT if insp.status == 'Open' else _GREEN_LIGHT if insp.status == 'Closed' else _YELLOW
            vals = [str(insp.inspection_date), insp.finding[:100], insp.category,
                    insp.location, insp.status, str(insp.due_date)]
            for col, val in enumerate(vals, 1):
                _cell(ws, r, col, val, fill=bg if col == 5 else None,
                      align='center' if col in (1, 5, 6) else 'left')
            r += 1
    r += 1

    # Incidents
    _section(ws, r, 'INCIDENT REPORTS', 6); r += 1
    _hdr(ws, r, ['Date', 'Type', 'Location', 'Description', 'Status', 'Injured Person']); r += 1
    incidents = IncidentReport.objects.filter(
        project=project, incident_date__range=(week_start, week_end)
    ).order_by('incident_date')
    if not incidents:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _cell(ws, r, 1, 'No incidents reported this week.', fill=_GREEN_LIGHT)
        r += 1
    else:
        for inc in incidents:
            vals = [str(inc.incident_date), inc.get_type_display(), inc.location,
                    inc.description[:100], inc.status, inc.injured_person or '—']
            for col, val in enumerate(vals, 1):
                _cell(ws, r, col, val, fill=_RED_LIGHT if col == 2 else None,
                      align='center' if col in (1, 2, 5) else 'left')
            r += 1


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_weekly_excel(project, week_start):
    """
    Return io.BytesIO containing the weekly report workbook.
    week_start: a datetime.date that is a Monday.
    """
    from apps.daily_reports.models import DailyReport

    week_end = week_start + timedelta(days=6)
    week_dates = _week_dates(week_start)

    reports = list(
        DailyReport.objects.filter(
            project=project,
            report_date__range=(week_start, week_end),
        ).select_related('prepared_by', 'checked_by', 'approved_by')
        .prefetch_related('activities__work_area', 'manpower_records__category', 'equipment_records__equipment')
        .order_by('report_date')
    )

    wb = openpyxl.Workbook()
    _build_summary(wb, project, week_start, reports, week_dates)
    _build_activities(wb, project, reports, week_dates)
    _build_manpower(wb, project, week_start, week_dates)
    _build_equipment(wb, project, week_start, week_dates)
    _build_safety(wb, project, week_start, week_dates)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
