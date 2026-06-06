import io
from datetime import date, timedelta
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import View
from apps.daily_reports.models import DailyReport
from apps.materials.models import MaterialDelivery, Material
from apps.boq.models import BOQItem
from apps.safety.models import SafetyInspection
from apps.projects.models import Project


class DailyReportPDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from apps.reports.pdf_generator import generate_daily_report_pdf
        report = get_object_or_404(DailyReport, pk=pk)
        buffer = generate_daily_report_pdf(report)
        filename = f"DailyReport_{report.project.contract_no}_{report.report_date}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class MaterialDeliveryExcelView(LoginRequiredMixin, View):
    def get(self, request, project_pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        project = get_object_or_404(Project, pk=project_pk)
        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')

        materials = Material.objects.filter(
            id__in=MaterialDelivery.objects.filter(project=project).values_list('material_id', flat=True)
        )

        for material in materials:
            ws = wb.create_sheet(title=material.name[:31])
            headers = ['Date', 'Delivery Note', 'Truck No', 'Source', 'Quantity', 'Unit', 'Unit Price', 'Total Amount', 'Remarks']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            deliveries = MaterialDelivery.objects.filter(project=project, material=material).order_by('delivery_date')
            for row_num, d in enumerate(deliveries, 2):
                ws.cell(row=row_num, column=1, value=str(d.delivery_date))
                ws.cell(row=row_num, column=2, value=d.delivery_note_no)
                ws.cell(row=row_num, column=3, value=d.truck_no)
                ws.cell(row=row_num, column=4, value=d.source)
                ws.cell(row=row_num, column=5, value=float(d.quantity))
                ws.cell(row=row_num, column=6, value=material.unit)
                ws.cell(row=row_num, column=7, value=float(d.unit_price) if d.unit_price else None)
                ws.cell(row=row_num, column=8, value=float(d.total_amount) if d.total_amount else None)
                ws.cell(row=row_num, column=9, value=d.remarks)

        summary = wb.active
        summary.title = 'Summary'
        sum_headers = ['Material', 'Unit', 'Total Delivered', 'Unit Price (avg)', 'Total Value']
        for col, h in enumerate(sum_headers, 1):
            cell = summary.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row_num, mat in enumerate(materials, 2):
            from apps.materials.models import get_stock_balance
            bal = get_stock_balance(project, mat)
            deliveries = MaterialDelivery.objects.filter(project=project, material=mat)
            avg_price = None
            if deliveries.filter(unit_price__isnull=False).exists():
                total_val = sum(float(d.total_amount) for d in deliveries if d.total_amount)
                avg_price = total_val / float(bal['delivered']) if bal['delivered'] else None
            summary.cell(row=row_num, column=1, value=mat.name)
            summary.cell(row=row_num, column=2, value=mat.unit)
            summary.cell(row=row_num, column=3, value=float(bal['delivered']))
            summary.cell(row=row_num, column=4, value=round(avg_price, 2) if avg_price else None)
            total_val = sum(float(d.total_amount) for d in deliveries if d.total_amount)
            summary.cell(row=row_num, column=5, value=total_val)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"MaterialDelivery_{project.contract_no}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BOQProgressExcelView(LoginRequiredMixin, View):
    def get(self, request, project_pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        project = get_object_or_404(Project, pk=project_pk)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'BOQ Progress'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')

        headers = ['Item No', 'Description', 'Unit', 'Contract Qty', 'Cumulative Qty',
                   'Remaining Qty', '% Complete', 'Contract Amount (THB)', 'Earned Value (THB)']
        col_widths = [10, 40, 8, 14, 14, 14, 12, 20, 20]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        items = BOQItem.objects.filter(project=project, item_type='item')
        total_contract = 0
        total_earned = 0
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item.item_no)
            ws.cell(row=row_num, column=2, value=item.description)
            ws.cell(row=row_num, column=3, value=item.unit)
            ws.cell(row=row_num, column=4, value=float(item.contract_quantity))
            ws.cell(row=row_num, column=5, value=float(item.cumulative_quantity))
            ws.cell(row=row_num, column=6, value=float(item.remaining_quantity))
            ws.cell(row=row_num, column=7, value=round(float(item.percent_complete), 2))
            ws.cell(row=row_num, column=8, value=float(item.contract_amount))
            ws.cell(row=row_num, column=9, value=float(item.earned_value))
            total_contract += float(item.contract_amount)
            total_earned += float(item.earned_value)
            if row_num % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=row_num, column=col).fill = PatternFill(
                        start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')

        total_row = len(items) + 2
        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=8, value=total_contract)
        ws.cell(row=total_row, column=9, value=total_earned)
        for col in range(1, 10):
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"BOQProgress_{project.contract_no}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PMCDailyReportPDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from apps.reports.pmc_pdf_generator import generate_pmc_pdf
        from apps.reports.validators import validate_daily_report_export
        report = get_object_or_404(
            DailyReport.objects.select_related('project', 'prepared_by', 'checked_by', 'approved_by'),
            pk=pk,
        )
        errors = validate_daily_report_export(report)
        hard_errors = [e for e in errors if not e.startswith('WARNING:')]
        if hard_errors:
            for msg in hard_errors:
                messages.error(request, msg)
            return redirect('daily_reports:detail', pk=pk)
        buf      = generate_pmc_pdf(report)
        filename = f"{report.project.contract_no}-DR-{report.report_date:%Y%m%d}-Rev0.pdf"
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PMCDailyReportExcelView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from apps.reports.pmc_excel_generator import generate_pmc_excel
        from apps.reports.validators import validate_daily_report_export
        report = get_object_or_404(
            DailyReport.objects.select_related('project', 'prepared_by', 'checked_by', 'approved_by'),
            pk=pk,
        )
        errors = validate_daily_report_export(report)
        hard_errors = [e for e in errors if not e.startswith('WARNING:')]
        if hard_errors:
            for msg in hard_errors:
                messages.error(request, msg)
            return redirect('daily_reports:detail', pk=pk)
        buf      = generate_pmc_excel(report)
        filename = f"{report.project.contract_no}-DR-{report.report_date:%Y%m%d}-Rev0.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class WeeklyReportView(LoginRequiredMixin, View):
    """Show form to pick a week, then download the weekly Excel report."""

    def _monday(self, d):
        return d - timedelta(days=d.weekday())

    def get(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        week_start_str = request.GET.get('week_start')
        if week_start_str:
            try:
                from datetime import datetime
                week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
                week_start = self._monday(week_start)
            except ValueError:
                week_start = self._monday(date.today())
        else:
            week_start = self._monday(date.today())
        return render(request, 'reports/weekly_report.html', {
            'project': project,
            'week_start': week_start,
            'week_end': week_start + timedelta(days=6),
        })

    def post(self, request, project_pk):
        project = get_object_or_404(Project, pk=project_pk)
        week_start_str = request.POST.get('week_start')
        try:
            from datetime import datetime
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
            week_start = self._monday(week_start)
        except (ValueError, TypeError):
            messages.error(request, 'Invalid week date.')
            return redirect('reports:weekly_report', project_pk=project_pk)

        from apps.reports.weekly_excel_generator import generate_weekly_excel
        buf = generate_weekly_excel(project, week_start)
        week_end = week_start + timedelta(days=6)
        filename = f"WeeklyReport_{project.contract_no}_{week_start}_{week_end}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class SafetyObservationExcelView(LoginRequiredMixin, View):
    def get(self, request, project_pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        project = get_object_or_404(Project, pk=project_pk)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Safety Observations'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='922B21', end_color='922B21', fill_type='solid')

        headers = ['Date', 'Location', 'Category', 'Finding', 'Action Required',
                   'Responsible', 'Due Date', 'Status', 'Close Date']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        inspections = SafetyInspection.objects.filter(project=project).order_by('-inspection_date')
        for row_num, insp in enumerate(inspections, 2):
            ws.cell(row=row_num, column=1, value=str(insp.inspection_date))
            ws.cell(row=row_num, column=2, value=insp.location)
            ws.cell(row=row_num, column=3, value=insp.category)
            ws.cell(row=row_num, column=4, value=insp.finding)
            ws.cell(row=row_num, column=5, value=insp.action_required)
            ws.cell(row=row_num, column=6, value=insp.responsible_person)
            ws.cell(row=row_num, column=7, value=str(insp.due_date))
            ws.cell(row=row_num, column=8, value=insp.status)
            ws.cell(row=row_num, column=9, value=str(insp.close_date) if insp.close_date else '')

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"SafetyObservations_{project.contract_no}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
