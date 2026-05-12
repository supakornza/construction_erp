"""
PMC-compliant Daily Site Report Excel generator.

Sheet 1 "Report View"  – formatted, print-ready (mirrors PDF sections)
Sheet 2 "Raw Data"     – flat table, no merged cells, Power BI / automation feed

Entry point: generate_pmc_excel(report) -> io.BytesIO
"""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────

_NAVY        = '1B2A4A'
_NAVY_LIGHT  = 'DCE8F5'
_WARN_YELLOW = 'FFF3CD'
_ERROR_RED   = 'F8D7DA'
_OK_GREEN    = 'D4EDDA'
_GREY_LIGHT  = 'F5F7FA'
_GREY_TEXT   = '6C757D'
_WHITE       = 'FFFFFF'

# ── Style factories ───────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=10, color='000000', italic=False):
    return Font(name='Arial', size=size, bold=bold, color=color, italic=italic)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border():
    s = Side(style='thin', color='B0BEC5')
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header_row(ws, row_num, values, col_widths=None):
    """Write a navy header row with white bold text."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font   = _font(bold=True, size=9, color=_WHITE)
        cell.fill   = _fill(_NAVY)
        cell.alignment = _align('center', 'center', wrap=True)
        cell.border = _thin_border()
    if col_widths:
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def _apply_data_cell(cell, value='', bold=False, fill_hex=None, align_h='left', wrap=True):
    cell.value     = value
    cell.font      = _font(bold=bold, size=9)
    cell.alignment = _align(align_h, 'center', wrap)
    cell.border    = _thin_border()
    if fill_hex:
        cell.fill = _fill(fill_hex)


def _section_title(ws, row_num, text, ncols):
    """Merge a row as a navy section bar across ncols columns."""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=ncols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font      = _font(bold=True, size=10, color=_WHITE)
    cell.fill      = _fill(_NAVY)
    cell.alignment = _align('left', 'center')
    ws.row_dimensions[row_num].height = 16


# ── Computed helpers ──────────────────────────────────────────────────────────

_COMPLETED_UNITS = {'ton', 'tonnes', 't', 'm3', 'm³', 'cum', 'm2', 'm²', 'm', 'kg', 'lm', 'pcs', 'no'}


def _activity_status(act):
    unit = (act.unit or '').lower().strip()
    qty  = act.quantity
    pct  = act.percent_complete
    if qty is not None and qty > Decimal('0') and unit in _COMPLETED_UNITS:
        return 'Completed', _OK_GREEN
    if pct is not None and pct > Decimal('0'):
        return 'In Progress', _WARN_YELLOW
    return 'Not Started', None


def _cumulative_qty(report, act):
    from django.db.models import Sum
    from apps.daily_reports.models import DailyWorkActivity
    result = DailyWorkActivity.objects.filter(
        report__project=report.project,
        report__report_date__lte=report.report_date,
        work_area=act.work_area,
        description=act.description,
    ).aggregate(total=Sum('quantity'))['total']
    return result or Decimal('0')


def _fmt(val):
    if val is None:
        return '—'
    try:
        d = Decimal(str(val))
        return float(d)
    except InvalidOperation:
        return str(val)


def _sig_name(user):
    if not user:
        return ''
    return user.get_full_name() or user.username


def _sig_role(user):
    return getattr(user, 'role', '') or ''


# ── Sheet 1: Report View ──────────────────────────────────────────────────────

def _build_report_view(wb, report):
    ws = wb.active
    ws.title = 'Report View'
    ws.sheet_properties.tabColor = _NAVY

    proj      = report.project
    report_no = f"{proj.contract_no}-DR-{report.report_date:%Y%m%d}"
    start_date = getattr(proj, 'start_date', None)
    day_count  = ((report.report_date - start_date).days + 1) if start_date else '—'
    status_lbl = 'FINAL' if report.status == 'Approved' else 'DRAFT'
    NCOLS      = 9

    # Page setup
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage  = True
    ws.print_title_rows      = '1:3'

    # Column widths (A–I)
    col_ws = [4, 12, 16, 10, 10, 12, 10, 12, 20]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value='PMC DAILY SITE REPORT')
    c.font      = _font(bold=True, size=14, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 22
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=proj.project_name)
    c.font      = _font(bold=True, size=11, color=_NAVY)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[r].height = 18
    r += 1
    r += 1  # spacer row

    # ── S1: Document Header ───────────────────────────────────────────────────
    _section_title(ws, r, '1.  DOCUMENT HEADER', NCOLS); r += 1

    info_pairs = [
        ('Project',      proj.project_name,        'Contract No', proj.contract_no),
        ('Report No',    report_no,                 'Report Date', str(report.report_date)),
        ('Day Count',    f'Day {day_count}' if start_date else '—', 'Status', status_lbl),
        ('Weather AM',   report.weather_morning,    'Weather PM',  report.weather_afternoon),
        ('Sea Condition','N/A',                     'Distribution','PMC / Owner / Contractor'),
        ('Contractor',   proj.contractor,           'Owner',       proj.owner),
    ]
    for lbl1, val1, lbl2, val2 in info_pairs:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        for col, (txt, is_label) in enumerate(
            [(lbl1, True), (val1, False), (lbl2, True), (val2, False)],
            start=0,
        ):
            real_col = [1, 2, 5, 6][col]
            cell = ws.cell(row=r, column=real_col, value=txt)
            cell.font      = _font(bold=is_label, size=9,
                                   color=_NAVY if is_label else '000000')
            cell.fill      = _fill(_NAVY_LIGHT) if is_label else _fill(_WHITE)
            cell.alignment = _align('left', 'center')
            cell.border    = _thin_border()
        r += 1
    r += 1

    # ── S2: Manpower ──────────────────────────────────────────────────────────
    _section_title(ws, r, '2.  MANPOWER', NCOLS); r += 1
    mp_headers = ['Category', 'Company', 'Planned', 'Actual', 'Remarks']
    mp_widths  = None
    _apply_header_row(ws, r, mp_headers); r += 1
    mp_records = report.manpower_records.select_related('category').all()
    total_actual = 0
    for rec in mp_records:
        total_actual += rec.quantity
        row_vals = [rec.category.get_name_display(), rec.company, '—', rec.quantity, rec.remarks]
        for col, val in enumerate(row_vals, 1):
            _apply_data_cell(ws.cell(row=r, column=col), val)
        r += 1
    if not mp_records:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        _apply_data_cell(ws.cell(row=r, column=1), 'No manpower records.', fill_hex=_GREY_LIGHT)
        r += 1
    # Total row
    for col, val in enumerate(['TOTAL', '', '—', total_actual, ''], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font   = _font(bold=True, size=9)
        c.fill   = _fill(_NAVY_LIGHT)
        c.border = _thin_border()
        c.alignment = _align('left', 'center')
    r += 2

    # ── S3: Equipment ─────────────────────────────────────────────────────────
    _section_title(ws, r, '3.  EQUIPMENT', NCOLS); r += 1
    eq_headers = ['Equipment', 'Status', 'Working Hrs', 'Idle Hrs', 'Utilization %', 'Remarks']
    _apply_header_row(ws, r, eq_headers); r += 1
    eq_records = report.equipment_records.select_related('equipment').all()
    for rec in eq_records:
        wh   = rec.working_hours or Decimal('0')
        idle = max(Decimal('0'), Decimal('10') - wh)
        util = round(float(wh) / 10.0 * 100, 1)
        bg   = _WARN_YELLOW if rec.status in ('Standby', 'Breakdown') else None
        row_vals = [rec.equipment.name, rec.status, float(wh), float(idle), util, rec.remarks]
        for col, val in enumerate(row_vals, 1):
            _apply_data_cell(ws.cell(row=r, column=col), val, fill_hex=bg)
        r += 1
    if not eq_records:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _apply_data_cell(ws.cell(row=r, column=1), 'No equipment records.', fill_hex=_GREY_LIGHT)
        r += 1
    r += 1

    # ── S4: Daily Work Activities ─────────────────────────────────────────────
    _section_title(ws, r, '4.  DAILY WORK ACTIVITIES', NCOLS); r += 1
    act_headers = ['No', 'Activity', 'Location', 'Unit', 'Today Qty',
                   'Cumul. Qty', '% Complete', 'Status', 'Remarks']
    _apply_header_row(ws, r, act_headers); r += 1
    activities = report.activities.select_related('work_area').all()
    if not activities:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        _apply_data_cell(ws.cell(row=r, column=1), 'No activities recorded.', fill_hex=_ERROR_RED)
        r += 1
    else:
        for i, act in enumerate(activities, 1):
            cumul = _cumulative_qty(report, act)
            status, bg = _activity_status(act)
            loc = str(act.work_area) if act.work_area else '—'
            row_vals = [i, act.description, loc, act.unit or '—',
                        _fmt(act.quantity), _fmt(cumul),
                        float(act.percent_complete) if act.percent_complete else '—',
                        status, act.remarks]
            for col, val in enumerate(row_vals, 1):
                fill = bg if col == 8 else None
                _apply_data_cell(ws.cell(row=r, column=col), val, fill_hex=fill)
            r += 1
    r += 1

    # ── S5: Problems / Issues / Safety ────────────────────────────────────────
    _section_title(ws, r, '5.  PROBLEMS / ISSUES / SAFETY', NCOLS); r += 1
    all_problems  = list(report.problems.all())
    tech_problems = [p for p in all_problems if p.category != 'Safety']
    safety_probs  = [p for p in all_problems if p.category == 'Safety']
    prob_headers  = ['Category', 'Description', 'Impact', 'Action', 'Status', '', '', '', '']

    ws.cell(row=r, column=1, value='Technical Issues').font = _font(bold=True, size=9, color=_NAVY)
    r += 1
    _apply_header_row(ws, r, prob_headers); r += 1
    if tech_problems:
        for p in tech_problems:
            for col, val in enumerate([p.category, p.description, p.impact,
                                       p.corrective_action, p.status, '', '', '', ''], 1):
                _apply_data_cell(ws.cell(row=r, column=col), val)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1, value='NIL — No issues reported')
        c.font = _font(italic=True, size=9, color=_GREY_TEXT)
        c.fill = _fill(_GREY_LIGHT)
        c.border = _thin_border()
        r += 1

    ws.cell(row=r, column=1, value='Safety').font = _font(bold=True, size=9, color=_NAVY)
    r += 1
    _apply_header_row(ws, r, prob_headers); r += 1
    if safety_probs:
        for p in safety_probs:
            for col, val in enumerate(['Safety', p.description, p.impact,
                                       p.corrective_action, p.status, '', '', '', ''], 1):
                _apply_data_cell(ws.cell(row=r, column=col), val)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1, value='NIL — No safety issues reported')
        c.font = _font(italic=True, size=9, color=_GREY_TEXT)
        c.fill = _fill(_GREY_LIGHT)
        c.border = _thin_border()
        r += 1

    # Safety indicators
    for col, (lbl, val) in enumerate(
        [('Near Miss', '—'), ('Toolbox Talk', '—'), ('PTW Active', '—')], 1,
    ):
        lc = ws.cell(row=r, column=col * 3 - 2, value=lbl)
        lc.font = _font(bold=True, size=9); lc.fill = _fill(_NAVY_LIGHT); lc.border = _thin_border()
        vc = ws.cell(row=r, column=col * 3 - 1, value=val)
        vc.font = _font(size=9, color=_GREY_TEXT); vc.border = _thin_border()
    r += 2

    # ── S6: Next Day Plan ─────────────────────────────────────────────────────
    _section_title(ws, r, '6.  NEXT DAY PLAN (LOOKAHEAD)', NCOLS); r += 1
    la_headers = ['Activity', 'Location', 'Planned Qty', 'Equipment Required',
                  'Responsible', 'Planned Date', '', '', '']
    _apply_header_row(ws, r, la_headers); r += 1
    lookaheads = report.lookaheads.all()
    if not lookaheads:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(row=r, column=1, value='No next-day plan recorded.')
        c.font = _font(italic=True, size=9, color=_GREY_TEXT)
        c.fill = _fill(_GREY_LIGHT); c.border = _thin_border()
        r += 1
    else:
        for la in lookaheads:
            for col, val in enumerate([la.planned_activity, '—', '—', '—',
                                       la.responsible_person, str(la.planned_date),
                                       '', '', ''], 1):
                _apply_data_cell(ws.cell(row=r, column=col), val)
            r += 1
    r += 1

    # ── S7: Document Control Footer ───────────────────────────────────────────
    _section_title(ws, r, '7.  DOCUMENT CONTROL', NCOLS); r += 1
    sig_labels = ['Prepared By', 'Checked By', 'Approved By']
    sig_users  = [report.prepared_by, report.checked_by, report.approved_by]
    col_map    = [1, 4, 7]
    for label, user, col in zip(sig_labels, sig_users, col_map):
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
        c = ws.cell(row=r, column=col, value=label)
        c.font = _font(bold=True, size=9); c.fill = _fill(_NAVY_LIGHT)
        c.alignment = _align('center'); c.border = _thin_border()
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 2)
        nv = ws.cell(row=r + 1, column=col, value=_sig_name(user))
        nv.font = _font(bold=True, size=9); nv.alignment = _align('center'); nv.border = _thin_border()
        ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2, end_column=col + 2)
        rv = ws.cell(row=r + 2, column=col, value=_sig_role(user))
        rv.font = _font(size=9, color=_GREY_TEXT); rv.alignment = _align('center')
        rv.border = _thin_border()
    r += 4
    ws.cell(row=r, column=1, value=f'Issue Date: {report.report_date}').font = _font(size=8, color=_GREY_TEXT)
    ws.cell(row=r, column=4, value='Revision: Rev 0').font = _font(size=8, color=_GREY_TEXT)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=NCOLS)
    ws.cell(row=r, column=6,
            value=f'File: {proj.contract_no}-DR-{report.report_date:%Y%m%d}-Rev0.xlsx'
    ).font = _font(size=8, color=_GREY_TEXT)


# ── Sheet 2: Raw Data ─────────────────────────────────────────────────────────

def _build_raw_data(wb, report):
    ws = wb.create_sheet(title='Raw Data')
    ws.sheet_properties.tabColor = '1A5276'

    proj       = report.project
    report_no  = f"{proj.contract_no}-DR-{report.report_date:%Y%m%d}"
    start_date = getattr(proj, 'start_date', None)
    day_count  = ((report.report_date - start_date).days + 1) if start_date else None
    status_lbl = 'FINAL' if report.status == 'Approved' else 'DRAFT'

    headers = [
        'report_no', 'project_name', 'contract_no', 'report_date', 'day_count',
        'status', 'weather_am', 'weather_pm',
        'activity_no', 'activity_description', 'location', 'unit',
        'today_qty', 'cumulative_qty', 'pct_complete', 'activity_status', 'remarks',
    ]
    col_widths = [22, 30, 16, 12, 10, 10, 14, 14, 10, 40, 20, 8, 12, 14, 12, 14, 30]
    _apply_header_row(ws, 1, headers, col_widths)

    common = [
        report_no,
        proj.project_name,
        proj.contract_no,
        str(report.report_date),
        day_count,
        status_lbl,
        report.weather_morning,
        report.weather_afternoon,
    ]

    activities = report.activities.select_related('work_area').all()
    if not activities:
        row_vals = common + [None, '(no activities)', None, None, None, None, None, None]
        for col, val in enumerate(row_vals, 1):
            ws.cell(row=2, column=col, value=val).border = _thin_border()
    else:
        for i, (seq, act) in enumerate(enumerate(activities, 1), 2):
            cumul  = _cumulative_qty(report, act)
            status, _ = _activity_status(act)
            loc    = str(act.work_area) if act.work_area else None
            row_vals = common + [
                seq,
                act.description,
                loc,
                act.unit or None,
                float(act.quantity) if act.quantity is not None else None,
                float(cumul),
                float(act.percent_complete) if act.percent_complete else None,
                status,
                act.remarks or None,
            ]
            for col, val in enumerate(row_vals, 1):
                ws.cell(row=i, column=col, value=val).border = _thin_border()


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_pmc_excel(report):
    """Return io.BytesIO containing the PMC-compliant daily report workbook."""
    wb = openpyxl.Workbook()
    _build_report_view(wb, report)
    _build_raw_data(wb, report)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
