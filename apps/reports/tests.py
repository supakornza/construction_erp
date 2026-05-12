"""
Unit tests for PMC daily report validators and generator smoke tests.
"""
import io
from decimal import Decimal
from unittest.mock import MagicMock, patch, PropertyMock

from django.test import TestCase

from apps.reports.validators import validate_daily_report_export


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_report(
    status='Draft',
    approved_by=None,
    activities=None,
    lookaheads=None,
    eq_records=None,
    mp_records=None,
):
    """Build a minimal fake DailyReport-like object for validator tests."""

    def _qs(items):
        qs = MagicMock()
        qs.count.return_value = len(items)
        qs.all.return_value    = items
        qs.__iter__            = lambda self: iter(items)
        qs.select_related      = lambda *a: qs
        return qs

    report               = MagicMock()
    report.status        = status
    report.approved_by   = approved_by
    report.activities    = _qs(activities  or [])
    report.lookaheads    = _qs(lookaheads  or [])
    report.equipment_records = _qs(eq_records or [])
    report.manpower_records  = _qs(mp_records  or [])
    return report


def _make_eq(name, status, hours, remarks=''):
    eq  = MagicMock()
    eq.equipment.name = name
    eq.status         = status
    eq.working_hours  = Decimal(str(hours))
    eq.remarks        = remarks
    return eq


def _make_mp(company, qty):
    mp          = MagicMock()
    mp.company  = company
    mp.quantity = qty
    return mp


def _make_act(desc, qty=None, unit='ton', pct=None):
    act             = MagicMock()
    act.description = desc
    act.quantity    = Decimal(str(qty)) if qty is not None else None
    act.unit        = unit
    act.percent_complete = Decimal(str(pct)) if pct is not None else None
    return act


# ── TestValidateDailyReportExport ─────────────────────────────────────────────

class TestValidateDailyReportExport(TestCase):

    def test_valid_report_returns_no_errors(self):
        act    = _make_act('Rock Fill', qty=450)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la])
        errors = validate_daily_report_export(report)
        hard   = [e for e in errors if not e.startswith('WARNING:')]
        self.assertEqual(hard, [])

    def test_approved_status_missing_approver_is_error(self):
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(status='Approved', approved_by=None,
                              activities=[act], lookaheads=[la])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('Approved By' in e for e in errors))

    def test_approved_status_with_approver_is_ok(self):
        act      = _make_act('Rock Fill', qty=100)
        la       = MagicMock()
        approver = MagicMock()
        report   = _make_report(status='Approved', approved_by=approver,
                                activities=[act], lookaheads=[la])
        errors   = validate_daily_report_export(report)
        hard     = [e for e in errors if not e.startswith('WARNING:')]
        self.assertEqual(hard, [])

    def test_standby_equipment_no_remarks_is_error(self):
        eq     = _make_eq('Wheel Loader', 'Standby', 0, remarks='')
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la], eq_records=[eq])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('Wheel Loader' in e and 'Standby' in e for e in errors))

    def test_standby_equipment_with_remarks_is_ok(self):
        eq     = _make_eq('Wheel Loader', 'Standby', 0, remarks='Hydraulic leak under repair')
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la], eq_records=[eq])
        errors = validate_daily_report_export(report)
        hard   = [e for e in errors if not e.startswith('WARNING:')]
        self.assertEqual(hard, [])

    def test_breakdown_equipment_no_remarks_is_error(self):
        eq     = _make_eq('Crane 50T', 'Breakdown', 2, remarks='')
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la], eq_records=[eq])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('Crane 50T' in e and 'Breakdown' in e for e in errors))

    def test_zero_activities_is_error(self):
        la     = MagicMock()
        report = _make_report(activities=[], lookaheads=[la])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('0 rows' in e for e in errors))

    def test_zero_lookaheads_is_warning_not_error(self):
        act    = _make_act('Rock Fill', qty=100)
        report = _make_report(activities=[act], lookaheads=[])
        errors = validate_daily_report_export(report)
        hard   = [e for e in errors if not e.startswith('WARNING:')]
        warns  = [e for e in errors if e.startswith('WARNING:')]
        self.assertEqual(hard, [])
        self.assertTrue(len(warns) > 0)

    def test_negative_manpower_is_error(self):
        mp     = _make_mp('ITD', -5)
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la], mp_records=[mp])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('negative' in e.lower() for e in errors))

    def test_negative_equipment_hours_is_error(self):
        eq     = _make_eq('Dump Truck', 'Working', -1)
        act    = _make_act('Rock Fill', qty=100)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la], eq_records=[eq])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('negative' in e.lower() for e in errors))

    def test_negative_activity_quantity_is_error(self):
        act    = _make_act('Rock Fill', qty=-10)
        la     = MagicMock()
        report = _make_report(activities=[act], lookaheads=[la])
        errors = validate_daily_report_export(report)
        self.assertTrue(any('negative' in e.lower() for e in errors))


# ── TestPMCPdfGenerator (smoke test) ─────────────────────────────────────────

class TestPMCPdfGeneratorSmoke(TestCase):

    def _make_full_report(self):
        project = MagicMock()
        project.project_name = 'Port Expansion Phase 1'
        project.contract_no  = 'CON-2024-001'
        project.contractor   = 'Test Contractor Co.'
        project.owner        = 'Port Authority'
        project.start_date   = None

        act       = MagicMock()
        act.work_area    = None
        act.description  = 'Rock Fill Installation Zone A'
        act.quantity     = Decimal('450')
        act.unit         = 'ton'
        act.percent_complete = Decimal('34')
        act.remarks      = ''

        la = MagicMock()
        la.planned_activity  = 'Continue rock fill Zone A'
        la.planned_date      = '2026-05-13'
        la.responsible_person= 'Site Manager'

        eq = MagicMock()
        eq.equipment.name = 'Crane 50T'
        eq.status         = 'Working'
        eq.working_hours  = Decimal('8.5')
        eq.remarks        = ''

        def _qs(items):
            qs = MagicMock()
            qs.count.return_value = len(items)
            qs.all.return_value   = items
            qs.__iter__           = lambda s: iter(items)
            qs.select_related     = lambda *a: qs
            return qs

        report = MagicMock()
        report.project          = project
        report.report_date      = __import__('datetime').date(2026, 5, 12)
        report.status           = 'Draft'
        report.weather_morning  = 'Sunny'
        report.weather_afternoon= 'Sunny'
        report.remarks          = ''
        report.prepared_by      = None
        report.checked_by       = None
        report.approved_by      = None
        report.activities       = _qs([act])
        report.lookaheads       = _qs([la])
        report.equipment_records= _qs([eq])
        report.manpower_records = _qs([])
        report.problems         = _qs([])
        return report

    @patch('apps.daily_reports.models.DailyWorkActivity.objects')
    def test_generates_nonempty_pdf(self, mock_mgr):
        mock_mgr.filter.return_value.aggregate.return_value = {'total': Decimal('12500')}
        from apps.reports.pmc_pdf_generator import generate_pmc_pdf
        buf = generate_pmc_pdf(self._make_full_report())
        self.assertIsInstance(buf, io.BytesIO)
        self.assertGreater(len(buf.getvalue()), 0)
        self.assertTrue(buf.getvalue().startswith(b'%PDF'))


# ── TestPMCExcelGenerator (smoke test) ────────────────────────────────────────

class TestPMCExcelGeneratorSmoke(TestCase):

    def _make_full_report(self):
        project = MagicMock()
        project.project_name = 'Port Expansion Phase 1'
        project.contract_no  = 'CON-2024-001'
        project.contractor   = 'Test Contractor Co.'
        project.owner        = 'Port Authority'
        project.start_date   = None

        act       = MagicMock()
        act.work_area    = None
        act.description  = 'Rock Fill Installation Zone A'
        act.quantity     = Decimal('450')
        act.unit         = 'ton'
        act.percent_complete = Decimal('34')
        act.remarks      = ''

        la = MagicMock()
        la.planned_activity   = 'Continue rock fill'
        la.planned_date       = '2026-05-13'
        la.responsible_person = 'Site Manager'

        def _qs(items):
            qs = MagicMock()
            qs.count.return_value = len(items)
            qs.all.return_value   = items
            qs.__iter__           = lambda s: iter(items)
            qs.select_related     = lambda *a: qs
            return qs

        report = MagicMock()
        report.project           = project
        report.report_date       = __import__('datetime').date(2026, 5, 12)
        report.status            = 'Draft'
        report.weather_morning   = 'Sunny'
        report.weather_afternoon = 'Sunny'
        report.remarks           = ''
        report.prepared_by       = None
        report.checked_by        = None
        report.approved_by       = None
        report.activities        = _qs([act])
        report.lookaheads        = _qs([la])
        report.equipment_records = _qs([])
        report.manpower_records  = _qs([])
        report.problems          = _qs([])
        return report

    @patch('apps.daily_reports.models.DailyWorkActivity.objects')
    def test_generates_workbook_with_two_sheets(self, mock_mgr):
        mock_mgr.filter.return_value.aggregate.return_value = {'total': Decimal('12500')}
        import openpyxl
        from apps.reports.pmc_excel_generator import generate_pmc_excel
        buf = generate_pmc_excel(self._make_full_report())
        self.assertIsInstance(buf, io.BytesIO)
        wb  = openpyxl.load_workbook(buf)
        self.assertIn('Report View', wb.sheetnames)
        self.assertIn('Raw Data',    wb.sheetnames)

    @patch('apps.daily_reports.models.DailyWorkActivity.objects')
    def test_raw_data_sheet_has_header_and_data_rows(self, mock_mgr):
        mock_mgr.filter.return_value.aggregate.return_value = {'total': Decimal('12500')}
        import openpyxl
        from apps.reports.pmc_excel_generator import generate_pmc_excel
        buf = generate_pmc_excel(self._make_full_report())
        wb  = openpyxl.load_workbook(buf)
        ws  = wb['Raw Data']
        self.assertEqual(ws.cell(row=1, column=1).value, 'report_no')
        self.assertGreaterEqual(ws.max_row, 2)
